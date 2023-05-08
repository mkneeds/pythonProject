import openpyxl

# загрузка данных из файла
workbook = openpyxl.load_workbook(filename='Kriptoplatforma_3.xlsx')
worksheet = workbook['Dynamics']

# ищем строку, которая начинается с "Количество юзеров"
for row in worksheet.iter_rows():
    for cell in row:
        if cell.value == 'Количество юзеров':
            start_row = cell.row
            break

# выбираем все ячейки, которые находятся правее найденной строки
result = []
for row in worksheet.iter_rows(min_row=start_row, min_col=worksheet.max_column):
    result.append([cell.value for cell in row])

for row in worksheet.iter_rows():
    for cell in row:
        print(cell.value, end='\t')
    print()

